#1
#v1.15

# -*- coding: utf-8 -*-
import openpyxl
import os
import xlrd
import subprocess
import re
import time
import sys
import logging
from pymorphy2 import MorphAnalyzer
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Отображение ошибок
def show_error_message(message):
    try:
        root = tk.Tk()
        root.withdraw()  # Скрываем главное окно Tkinter
        messagebox.showerror("Ошибка", message)
        root.destroy()
    except Exception as e:
        logging.error(f"Ошибка при отображении сообщения об ошибке: {str(e)}")

# Инициализируем морфологический анализатор
morph = MorphAnalyzer()

# Преобразует слово из родительного падежа в именительный падеж.
def convert_to_nominative(word):
    try:
        logging.info(f"Преобразуем слово '{word}' из родительного падежа в именительный падеж.")
        parsed_word = morph.parse(word)[0]
        if parsed_word:
            nominative_form = parsed_word.inflect({'nomn'}).word
            logging.info(f"Преобразованное слово: '{nominative_form.capitalize()}'")
            return nominative_form.capitalize()
        else:
            logging.error(f"Не удалось разобрать слово: '{word}'")
            return word
    except Exception as e:
        logging.error(f"Ошибка при преобразовании слова '{word}': {str(e)}")
        return word

# Удаление пробелов в начале и конце строки
def strip_spaces(value):
    if value and isinstance(value, str):
        return value.strip()
    return value

# Заменяет точку на запятую в числовых значениях
def replace_dot_with_comma(value):
    if isinstance(value, (int, float)):
        value_str = str(value)
        if ',' not in value_str:
            return value_str.replace('.', ',')
    elif isinstance(value, str) and '.' in value and value.replace('.', '').isdigit():
        return value.replace('.', ',')
    return str(value) if value is not None else value

# Извлекает дату из имени файла. Возвращает строку с датой или исходное имя, если дата не найдена.
def extract_date_from_filename(filename):
    match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    if match:
        return match.group(1)
    match = re.search(r'(\d{2}\.\d{2}\.\d{4})', filename)
    if match:
        return match.group(1)
    return filename

# Показывает диалог выбора файла из списка matches для образца obrt. Возвращает выбранную запись (словарь с полями dm, km, yx, id, file_name).
def ask_user_to_choose_file(matches, obrt):
    """
    Предлагает пользователю выбрать один из нескольких файлов через консоль.
    matches: список словарей с ключами dm, km, yx, id, file_name.
    obrt: имя образца (для информационных сообщений).
    Возвращает выбранный словарь.
    При ошибке или некорректном вводе автоматически выбирается второй вариант (индекс 1).
    """
    print(f"\nДля образца '{obrt}' найдено несколько файлов. Выберите нужный:")
    for idx, match in enumerate(matches, start=1):
        date_str = extract_date_from_filename(os.path.basename(match['file_name']))
        print(f"{idx}. {date_str} - {match['file_name']}")

    # Если вдруг меньше двух вариантов (не должно быть, но на всякий случай)
    if len(matches) < 2:
        logging.warning(f"Для образца {obrt} вариантов меньше двух, выбираем первый.")
        return matches[0]

    while True:
        try:
            choice = input("Введите номер варианта (или нажмите Enter для выбора второго по умолчанию): ").strip()
            if choice == "":
                logging.info("Пользователь пропустил ввод, выбираем второй вариант по умолчанию.")
                return matches[1]

            choice_num = int(choice)
            if 1 <= choice_num <= len(matches):
                logging.info(f"Пользователь выбрал вариант {choice_num}: {matches[choice_num-1]['file_name']}")
                return matches[choice_num-1]
            else:
                print(f"Ошибка: введите число от 1 до {len(matches)}.")
                # Не выходим из цикла, даём ещё попытку
        except ValueError:
            print("Ошибка: введите целое число.")
        except (KeyboardInterrupt, EOFError):
            # При прерывании (Ctrl+C) выбираем второй вариант и продолжаем работу
            logging.warning("Получен сигнал прерывания. Автоматически выбираем второй вариант.")
            return matches[1]
        except Exception as e:
            logging.error(f"Неожиданная ошибка при выборе: {e}. Выбираем второй вариант.")
            return matches[1]

# Функция для поиска значений РТ в папке
def find_values_in_excel_files(folder_path, obrt_values):
    start_time = time.time()
    logging.info(f"Начало поиска значений для ОБРТ: {obrt_values} в папке {folder_path}")

    default_values = {
        "dm": "ниже порога детекции",
        "km": "ниже порога детекции",
        "yx": "ниже порога детекции",
        "id": "-",
    }

    obrt_set = set(obrt_values)
    all_matches = {obrt: [] for obrt in obrt_values}
    files_processed = 0

    # Внутренняя функция для поиска индексов столбцов по заголовкам
    def _find_column_indices(sheet, file_ext):
        """
        Ищет индексы столбцов для заголовков 'Quantity' и 'Degradation Index Mean'.
        Возвращает dict {'qty': idx, 'deg': idx} с индексами (0-базовыми для xlrd,
        для openpyxl индексы конвертируются в 0-базу).
        Если заголовок не найден, соответствующий ключ получает значение None.
        """
        indices = {'qty': None, 'deg': None}
        search_rows = [8]  # сначала проверяем 8-ю строку
        # если в 8-й не найдено, расширяем поиск на первые 20 строк
        # (чтобы не сканировать весь лист)
        for r in range(1, 21):
            if r != 8:
                search_rows.append(r)

        # Определяем функцию для доступа к значению ячейки в зависимости от типа файла
        if file_ext == '.xlsx':
            # openpyxl: строки и колонки 1-базовые
            def cell_value(row_idx, col_idx):
                # row_idx и col_idx здесь 1-базовые
                cell = sheet.cell(row=row_idx, column=col_idx)
                return cell.value if cell else None
            # для преобразования в 0-базовый индекс вычтем 1 позже
        else:  # .xls
            # xlrd: строки и колонки 0-базовые
            def cell_value(row_idx, col_idx):
                # row_idx и col_idx здесь 1-базовые, приводим к 0-базе
                if row_idx-1 < sheet.nrows and col_idx-1 < sheet.ncols:
                    return sheet.cell_value(row_idx-1, col_idx-1)
                return None

        # Перебираем строки в порядке приоритета
        for row_num in search_rows:
            # Если для openpyxl строка выходит за пределы используемого диапазона,
            # просто пропускаем (sheet.max_row может быть меньше)
            # Для xlrd аналогично проверяем row_num-1 < nrows
            if file_ext == '.xlsx' and row_num > sheet.max_row:
                continue
            if file_ext == '.xls' and row_num-1 >= sheet.nrows:
                continue

            # Сканируем колонки с 1 по 20 (обычно достаточно)
            for col in range(1, 21):
                val = cell_value(row_num, col)
                if val is None:
                    continue
                val_str = str(val).strip()
                if val_str == '':
                    continue

                # Проверяем нужные заголовки
                if indices['qty'] is None and val_str == "Quantity":
                    indices['qty'] = col - 1  # переводим в 0-базу
                    logging.debug(f"Найден заголовок 'Quantity' в строке {row_num}, колонка {col}")
                    if row_num != 8:
                        logging.warning(f"Заголовок 'Quantity' найден в строке {row_num}, а не в 8-й. Используется колонка {col}.")
                if indices['deg'] is None and val_str == "Degradation Index Mean":
                    indices['deg'] = col - 1
                    logging.debug(f"Найден заголовок 'Degradation Index Mean' в строке {row_num}, колонка {col}")
                    if row_num != 8:
                        logging.warning(f"Заголовок 'Degradation Index Mean' найден в строке {row_num}, а не в 8-й. Используется колонка {col}.")

            # Если оба индекса уже найдены, прерываем поиск
            if indices['qty'] is not None and indices['deg'] is not None:
                break

        return indices

    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            file_path = os.path.join(root, file_name)

            if file_name.startswith('~$'):
                continue
            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                continue

            files_processed += 1
            logging.debug(f"Обработка файла #{files_processed}: {file_path}")

            file_data = {}
            file_ext = os.path.splitext(file_name)[1].lower()

            try:
                if file_ext == '.xlsx':
                    # Открываем без read_only, чтобы можно было искать заголовки и потом читать данные
                    workbook = openpyxl.load_workbook(file_path, read_only=False)
                    sheet = workbook.worksheets[0]
                    # Определяем индексы столбцов
                    col_indices = _find_column_indices(sheet, file_ext)
                    # Если не удалось найти нужные столбцы, пропускаем файл с ошибкой
                    if col_indices['qty'] is None:
                        logging.error(f"В файле {file_name} не найден столбец с заголовком 'Quantity'. Файл пропущен.")
                        workbook.close()
                        continue
                    if col_indices['deg'] is None:
                        logging.error(f"В файле {file_name} не найден столбец с заголовком 'Degradation Index Mean'. Файл пропущен.")
                        workbook.close()
                        continue

                    # Читаем данные со 2-й строки
                    rows = sheet.iter_rows(min_row=2, values_only=True)
                else:  # .xls
                    workbook = xlrd.open_workbook(file_path)
                    sheet = workbook.sheet_by_index(0)
                    col_indices = _find_column_indices(sheet, file_ext)
                    if col_indices['qty'] is None:
                        logging.error(f"В файле {file_name} не найден столбец с заголовком 'Quantity'. Файл пропущен.")
                        continue
                    if col_indices['deg'] is None:
                        logging.error(f"В файле {file_name} не найден столбец с заголовком 'Degradation Index Mean'. Файл пропущен.")
                        continue

                    # Для .xls собираем строки в список (как и раньше)
                    rows = []
                    for row_idx in range(1, sheet.nrows):
                        rows.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])

                # Логируем найденные индексы
                logging.info(f"Файл {file_name}: индексы столбцов Quantity={col_indices['qty']+1}, Degradation Index Mean={col_indices['deg']+1} (нумерация с 1)")

                # Обрабатываем строки
                for row in rows:
                    # Для .xlsx row - кортеж, для .xls - список, но обращаемся по индексу
                    if len(row) < 3:   # минимум должно быть 3 колонки (Sample Name, Target)
                        continue

                    # Используем индексы, но Sample Name и Target по-прежнему берем как 2-й и 3-й столбцы
                    sample_name = strip_spaces(str(row[1]))   # второй столбец (индекс 1)
                    if sample_name not in obrt_set:
                        continue

                    row_type = strip_spaces(str(row[2]))      # третий столбец (индекс 2)
                    if not row_type:
                        continue

                    if sample_name not in file_data:
                        file_data[sample_name] = default_values.copy()

                    def parse_value(value, is_id=False):
                        if value is None:
                            return None
                        value = strip_spaces(str(value))
                        if value.replace(".", "", 1).isdigit():
                            return round(float(value), 2 if is_id else 4)
                        return None

                    # Извлекаем значения Quantity и Degradation Index Mean по найденным индексам
                    if row_type == "T.Large Autosomal":
                        qty_val = row[col_indices['qty']] if len(row) > col_indices['qty'] else None
                        deg_val = row[col_indices['deg']] if len(row) > col_indices['deg'] else None

                        dm_value = parse_value(qty_val)
                        id_value = parse_value(deg_val, is_id=True)
                        if dm_value is not None:
                            file_data[sample_name]["dm"] = replace_dot_with_comma(dm_value)
                        if id_value is not None:
                            file_data[sample_name]["id"] = replace_dot_with_comma(id_value if id_value >= 1 else 0)

                    elif row_type == "T.Small Autosomal":
                        qty_val = row[col_indices['qty']] if len(row) > col_indices['qty'] else None
                        km_value = parse_value(qty_val)
                        if km_value is not None:
                            file_data[sample_name]["km"] = replace_dot_with_comma(km_value)

                    elif row_type == "T.Y":
                        qty_val = row[col_indices['qty']] if len(row) > col_indices['qty'] else None
                        yx_value = parse_value(qty_val)
                        if yx_value is not None:
                            file_data[sample_name]["yx"] = replace_dot_with_comma(yx_value)

                if file_ext == '.xlsx':
                    workbook.close()

                for sample_name, values in file_data.items():
                    match_entry = values.copy()
                    match_entry['file_name'] = file_path
                    all_matches[sample_name].append(match_entry)
                    logging.debug(f"Для образца {sample_name} добавлена запись из файла {file_path}")

            except Exception as e:
                logging.error(f"Ошибка при обработке файла {file_name}: {str(e)}")
                continue

    output = []
    found_status = []
    for obrt in obrt_values:
        matches = all_matches[obrt]
        if not matches:
            final_vals = default_values.copy()
            found_status.append(False)
            logging.info(f"Для образца {obrt} не найдено ни одного файла.")
        elif len(matches) == 1:
            final_vals = matches[0]
            found_status.append(True)
            logging.info(f"Для образца {obrt} найден один файл: {matches[0]['file_name']}")
        else:
            logging.info(f"Для образца {obrt} найдено {len(matches)} файлов. Требуется выбор пользователя.")
            try:
                chosen = ask_user_to_choose_file(matches, obrt)
                final_vals = chosen
                found_status.append(True)
                logging.info(f"Итоговый файл для образца {obrt}: {chosen['file_name']}")
            except Exception as e:
                logging.error(f"Ошибка при выборе файла для образца {obrt}: {str(e)}. Используется второй вариант.")
                # Если вдруг ошибка, берём второй (если есть), иначе первый
                final_vals = matches[1] if len(matches) > 1 else matches[0]
                found_status.append(True)
                logging.warning(f"Использован файл: {final_vals['file_name']}")

        output.extend([
            final_vals["dm"],
            final_vals["km"],
            final_vals["yx"],
            final_vals["id"],
        ])

    elapsed = time.time() - start_time
    logging.info(f"Поиск завершён. Обработано файлов: {files_processed}, время: {elapsed:.2f} сек")
    return (tuple(output), found_status)

# Функция для загрузки почтовых индексов из файла Excel
def load_postal_codes(file_path):
    postal_codes = {}
    try:
        logging.info(f"Загрузка почтовых индексов из файла: {file_path}")
        
        # Проверяем существование файла
        if not os.path.exists(file_path):
            logging.error(f"Файл с почтовыми индексами не найден: {file_path}")
            return postal_codes
            
        if file_path.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] and row[1]:
                    city = str(row[0]).strip().lower()
                    postal_code = str(row[1]).strip()
                    postal_codes[city] = postal_code
        else:  # .xls
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            for row_idx in range(1, sheet.nrows):
                city = str(sheet.cell_value(row_idx, 0)).strip().lower()
                postal_code = str(sheet.cell_value(row_idx, 1)).strip()
                postal_codes[city] = postal_code
                
        logging.info(f"Загружено {len(postal_codes)} почтовых индексов")
        return postal_codes
    except Exception as e:
        logging.error(f"Ошибка при загрузке почтовых индексов: {str(e)}")
        return {}

# Функция для извлечения населённого пункта из адреса
def extract_city_from_address(address):
    try:
        logging.info(f"Извлечение населённого пункта из адреса: '{address}'")
        address = str(address).lower()
        
        # Паттерны для поиска населённого пункта
        patterns = [
            "г. ",
            "город ",
            "г ",
            "пос. ",
            "пос ",
            "п. ",
            "п ",
            "пгт. ",
            "пгт ",
            "село ",
            "с. ",
            "с ",
            "дер. ",
            "д. ",
            "д "
        ]
        
        # Ищем начало названия населённого пункта
        start_idx = -1
        for pattern in patterns:
            idx = address.find(pattern)
            if idx != -1:
                start_idx = idx + len(pattern)
                break
        
        if start_idx == -1:
            logging.warning("Шаблон населённого пункта не найден в адресе")
            return None
        
        # Извлекаем название населённого пункта
        city_end = address.find(",", start_idx)
        if city_end == -1:
            city = address[start_idx:]
        else:
            city = address[start_idx:city_end]
        
        city = city.strip()
        logging.info(f"Извлеченный населённый пункт: '{city}'")
        return city
    except Exception as e:
        logging.error(f"Ошибка при извлечении населённого пункта: {str(e)}")
        return None

# Главная функция
def main():
    # Словарь с маппингом значений ТЕР -> ТЕРИ/ТЕР1/ТЕР2
    ter_mapping = {
        "Тюменской области": {
            "ТЕРИ": "Тюменская область",
            "ТЕР1": "Поштаренко Д.С.",
            "ТЕР2": "625025, г. Тюмень, Приисковый пер., д. 12"
        },
        "Ханты-Мансийскому АО - Югре": {
            "ТЕРИ": "ХМАО-Югра",
            "ТЕР1": "Илыку И.А.",
            "ТЕР2": "628011, ХМАО-Югра, г. Ханты-Мансийск, ул. Энгельса, д. 45"
        },
        "Ямало-Ненецкому АО": {
            "ТЕРИ": "ЯНАО",
            "ТЕР1": "Казановской Н.В.",
            "ТЕР2": "629008, ЯНАО, г. Салехард, ул. Республики, д. 73"
        }
    }

    try:
        # Получаем текущую директорию
        current_directory = os.path.dirname(os.path.abspath(__file__))
        logging.info(f"Текущая директория: {current_directory}")
        
        # Находим все файлы Excel в текущей директории
        excel_files = [f for f in os.listdir(current_directory) if f.endswith('.xlsx')]
        logging.info(f"Найденные файлы Excel: {excel_files}")
        
        # Проверяем, что файл Excel существует и единственный
        if not excel_files:
            raise FileNotFoundError("В текущей директории не найдено файлов Excel.")
        if len(excel_files) != 1:
            raise FileNotFoundError("В текущей директории должен быть только один файл Excel.")
        logging.info(f"Файл Excel найден: {excel_files[0]}")
        
        # Полный путь к файлу Excel
        excel_file_path = os.path.join(current_directory, excel_files[0])
        logging.info(f"Полный путь к файлу Excel: {excel_file_path}")
        
        # Открываем файл Excel
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active
            logging.info(f"Файл Excel открыт успешно.")
        except Exception as e:
            raise Exception(f"Ошибка при открытии файла Excel: {str(e)}")
        
        # Проверка условия для кода 1
        code_1_value = None
        obrt_values = []
        obrt2_value = None
        obrt3_value = None

        # Собираем значения для кода 1 и ОБРТ, ОБРТ2, ОБРТ3
        for row in sheet.iter_rows(min_row=2, max_col=3):
            code = row[0].value
            value = row[2].value
            if code == 1:
                code_1_value = str(value).strip() if value else None
            elif code == 'ОБРТ':
                obrt_value = str(value).strip() if value else None
                if obrt_value:
                    obrt_values.append(obrt_value)
            elif code == 'ОБРТ2':
                obrt2_value = str(value).strip() if value else None
                if obrt2_value:
                    obrt_values.append(obrt2_value)
            elif code == 'ОБРТ3':
                obrt3_value = str(value).strip() if value else None
                if obrt3_value:
                    obrt_values.append(obrt3_value)

        # Проверяем условие для завершения программы
        svo_condition = code_1_value in ["Кость СВО", "Кость СВО нет пригодной ДНК", "СВО_кость", "СВО_кость_нет результата"]

        # Поиск значения ТЕР
        ter_value = None
        for row in sheet.iter_rows(min_row=2, max_col=3):
            if row[0].value == "ТЕР" and row[2].value:
                ter_value = strip_spaces(str(row[2].value))
                logging.info(f"Найдено значение ТЕР: '{ter_value}'")
                break

        # Подготовка значений для ТЕРИ/ТЕР1/ТЕР2
        ter_results = {}
        if ter_value in ter_mapping:
            ter_results = ter_mapping[ter_value]
            logging.info(f"Подготовлены значения для ТЕРИ/ТЕР1/ТЕР2: {ter_results}")
        else:
            logging.warning(f"Значение ТЕР '{ter_value}' не соответствует ожидаемым условиям")

        # Проверка наличия значений ОБРТ
        if not any(obrt_values) and svo_condition:
            error_msg = "Ошибка: В графе с кодом 1 указано 'СВО_кость' или 'СВО_кость_нет результата', но значения ОБРТ отсутствуют"
            logging.error(error_msg)
            raise ValueError(error_msg)

        # Проверяем наличие значений в файлах Excel для всех указанных ОБРТ
        try:
            if obrt_values:
                found_results = find_values_in_excel_files(r"U:\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\АНАЛИЗАТОР\Работа\2-Реал-тайм\2026", obrt_values)
                found_values, found_samples = found_results
                
                # Проверка для ОБРТ
                if 'ОБРТ' in [code[0].value for code in sheet.iter_rows(min_row=2, max_col=1) if code[0].value]:
                    if not found_samples[0]:  # Проверяем, был ли найден Sample Name
                        error_msg = f"Значения для ОБРТ '{obrt_values[0]}' не найдены в файлах Excel"
                        logging.error(error_msg)
                        if svo_condition:
                            raise ValueError(error_msg)
                
                # Проверка для ОБРТ2
                if 'ОБРТ2' in [code[0].value for code in sheet.iter_rows(min_row=2, max_col=1) if code[0].value] and len(obrt_values) > 1:
                    if not found_samples[1]:  # Проверяем, был ли найден Sample Name
                        error_msg = f"Значения для ОБРТ2 '{obrt_values[1]}' не найдены в файлах Excel"
                        logging.error(error_msg)
                        if svo_condition:
                            raise ValueError(error_msg)
                
                # Проверка для ОБРТ3
                if 'ОБРТ3' in [code[0].value for code in sheet.iter_rows(min_row=2, max_col=1) if code[0].value] and len(obrt_values) > 2:
                    if not found_samples[2]:  # Проверяем, был ли найден Sample Name
                        error_msg = f"Значения для ОБРТ3 '{obrt_values[2]}' не найдены в файлах Excel"
                        logging.error(error_msg)
                        if svo_condition:
                            raise ValueError(error_msg)

        except Exception as e:
            if svo_condition:
                raise Exception(f"Ошибка при поиске значений для ОБРТ: {str(e)}")
            else:
                logging.error(f"Ошибка при поиске значений для ОБРТ: {str(e)}")
                    
        # Новые значения, которые нужно вставить в столбец 3
        new_values = {
            '1СТ': 'Бевза Алексей Леонидович',
            '2СТ': 'образование высшее медицинское, специальность – судебно-медицинская экспертиза,',
            '3СТ': 'стаж работы по специальности с 1.09.2004 г., высшая квалификационная категория,',
            '4СТ': 'кандидат медицинских наук, врач – судебно-медицинский эксперт',
            '5СТ': ' ',
            '6СТ': 'произвёл',
            '1ИЭ': 'А.Л. Бевза',
            '2ИЭ': '                                                     Бевза А.Л.',
            '3ИЭ': 'Бевза Алексей Леонидович',
            '4ИЭ': 'врач – судебно-медицинский эксперт ',
            '5ИЭ': 'отделения судебно-биологической',
            '6ИЭ': 'и генетической экспертизы, к.м.н.',
            '7ИЭ': 'тел. +7 (3452) 49-43-98 (доб. 1345)',
            2: 'U:\\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\\Беуза\\Экспертизы\\2026\\Уголовные',
            4: 'U:\\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\\Беуза\\ANALIZATOR',
            5: 'U:\\Архив\\Архив 2026',
            6: 'U:\\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\\Ростов СВО',
        }

        # Определяем значение для кода 3 на основе значения кода 1
        if code_1_value in ["СВО_Молов_образец_родственники", "СВО_Ростов_образец_родственники", 
                        "СВО_Молов_образец_прямая идентификация", "СВО_Ростов_образец_прямая идентификация", 
                        "СВО_Молов_образец_родственники_нет результата_RT", "СВО_Молов_образец_родственники_нет результата_форез"]:
            new_values[3] = "U:\\ШАБЛОНЫ\\Заключения\\СВО\\Образцы"
        elif code_1_value in ["СВО_кость", "СВО_кость_нет результата"]:
            new_values[3] = "U:\\ШАБЛОНЫ\\Заключения\\СВО\\Кости"
        elif code_1_value in ["ЭКЦ", "ЭКЦ_нет результата"] or (code_1_value and isinstance(code_1_value, str) and code_1_value.startswith("ЭКЦ_образец_")):
            new_values[3] = "U:\\ШАБЛОНЫ\\Заключения\\ЭКЦ"
        else:
            new_values[3] = "U:\\ШАБЛОНЫ\\Заключения\\СВО\\Образцы"  # значение по умолчанию, если ни одно условие не сработало

        logging.info(f"Новые значения для вставки: {new_values}")
        
        # СБОР ЗНАЧЕНИЙ ОБРТ, ОБРТ2, ОБРТ3 ДЛЯ ОДНОКРАТНОГО ПОИСКА
        obrt_search_values = []
        for r in sheet.iter_rows(min_row=2, max_col=3):
            if r[0].value in ['ОБРТ', 'ОБРТ2', 'ОБРТ3'] and r[2].value and str(r[2].value).strip():
                obrt_search_values.append(r[2].value)

        # Если есть что искать – выполняем поиск ОДИН РАЗ
        search_results = None
        if obrt_search_values:
            try:
                search_results = find_values_in_excel_files(r"U:\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\АНАЛИЗАТОР\Работа\2-Реал-тайм\2026", obrt_search_values)
                logging.info("Результаты поиска для всех ОБРТ получены однократно.")
                
                # Детальный вывод найденных значений
                found_values, found_samples = search_results
                num_obrt = len(obrt_search_values)
                if len(found_values) == 4:
                    dm, km, yx, id_ = found_values
                    logging.info(f"Найденные значения для {obrt_search_values[0]}: ДМ={dm}, КМ={km}, YХ={yx}, ИД={id_}")
                elif len(found_values) == 8:
                    dm1, km1, yx1, id1, dm2, km2, yx2, id2 = found_values
                    logging.info(f"Найденные значения для {obrt_search_values[0]}: ДМ={dm1}, КМ={km1}, YХ={yx1}, ИД={id1}")
                    logging.info(f"Найденные значения для {obrt_search_values[1]}: ДМ={dm2}, КМ={km2}, YХ={yx2}, ИД={id2}")
                elif len(found_values) == 12:
                    dm1, km1, yx1, id1, dm2, km2, yx2, id2, dm3, km3, yx3, id3 = found_values
                    logging.info(f"Найденные значения для {obrt_search_values[0]}: ДМ={dm1}, КМ={km1}, YХ={yx1}, ИД={id1}")
                    logging.info(f"Найденные значения для {obrt_search_values[1]}: ДМ={dm2}, КМ={km2}, YХ={yx2}, ИД={id2}")
                    logging.info(f"Найденные значения для {obrt_search_values[2]}: ДМ={dm3}, КМ={km3}, YХ={yx3}, ИД={id3}")
                else:
                    logging.warning(f"Неожиданная длина результатов поиска: {len(found_values)}")

                # Проверка на полное отсутствие данных (все значения по умолчанию и ни одного найденного образца)
                def is_default(val, field_type):
                    if field_type in ('dm','km','yx'):
                        return val == "ниже порога детекции"
                    else:  # id
                        return val == "-"

                all_default = True
                # Анализируем в зависимости от длины found_values
                if len(found_values) == 12:
                    vals = list(found_values)
                    field_types = ['dm','km','yx','id'] * 3
                    for i, val in enumerate(vals):
                        if not is_default(val, field_types[i % 4]):
                            all_default = False
                            break
                elif len(found_values) == 8:
                    vals = list(found_values)
                    field_types = ['dm','km','yx','id'] * 2
                    for i, val in enumerate(vals):
                        if not is_default(val, field_types[i % 4]):
                            all_default = False
                            break
                elif len(found_values) == 4:
                    vals = list(found_values)
                    field_types = ['dm','km','yx','id']
                    for i, val in enumerate(vals):
                        if not is_default(val, field_types[i]):
                            all_default = False
                            break
                else:
                    all_default = True  # неожиданная длина – считаем, что всё дефолтно

                if all_default and not any(found_samples):
                    error_msg = f"Значения для ОБРТ {obrt_search_values} не найдены в файлах Excel (все значения по умолчанию)."
                    logging.error(error_msg)
                    show_error_message(error_msg)
                    # search_results остаётся как есть (дефолтные значения) – они будут использованы в цикле

            except Exception as e:
                logging.error(f"Ошибка при поиске значений для ОБРТ: {str(e)}")
                show_error_message(f"Ошибка: {str(e)}")
                # search_results останется None – позже для всех кодов будет проставлено значение по умолчанию
                
        # Проходим по строкам и обновляем значения в столбец 3 для заданных кодов
        try:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                
                # Если код находится в new_values, обновляем значение
                if code in new_values:
                    sheet.cell(row=row[0].row, column=3).value = new_values[code]
                    logging.info(f"Обновлено значение для кода '{code}': '{new_values[code]}'")
                
                # Если код — ДМ, КМ, YХ, ИД, ДМ2, КМ2, YХ2, ИД2, ДМ3, КМ3, YХ3, ИД3, ищем значения в файлах Excel
                elif code in ['ДМ', 'КМ', 'YХ', 'ИД', 'ДМ2', 'КМ2', 'YХ2', 'ИД2', 'ДМ3', 'КМ3', 'YХ3', 'ИД3']:
                    # Используем ранее полученные результаты поиска
                    if search_results is None:
                        sheet.cell(row=row[0].row, column=3).value = "ниже порога детекции"
                        continue

                    found_values, found_samples = search_results

                    # Распаковываем значения в зависимости от количества найденных образцов
                    if len(found_values) == 12:
                        dm_value, km_value, yx_value, id_value, dm2_value, km2_value, yx2_value, id2_value, dm3_value, km3_value, yx3_value, id3_value = found_values
                    elif len(found_values) == 8:
                        dm_value, km_value, yx_value, id_value, dm2_value, km2_value, yx2_value, id2_value = found_values
                        dm3_value = km3_value = yx3_value = "ниже порога детекции"
                        id3_value = "-"
                    elif len(found_values) == 4:
                        dm_value, km_value, yx_value, id_value = found_values
                        dm2_value = km2_value = yx2_value = "ниже порога детекции"
                        id2_value = "-"
                        dm3_value = km3_value = yx3_value = "ниже порога детекции"
                        id3_value = "-"
                    else:
                        logging.error(f"Неожиданная длина found_values: {len(found_values)}")
                        sheet.cell(row=row[0].row, column=3).value = "ниже порога детекции"
                        continue

                    # Присваиваем значение в зависимости от конкретного кода
                    if code == 'ДМ':
                        sheet.cell(row=row[0].row, column=3).value = dm_value
                    elif code == 'КМ':
                        sheet.cell(row=row[0].row, column=3).value = km_value
                    elif code == 'YХ':
                        sheet.cell(row=row[0].row, column=3).value = yx_value
                    elif code == 'ИД':
                        # Проверяем условие для ИД
                        if dm_value == "ниже порога детекции" and km_value != "ниже порога детекции" and km_value != "-":
                            sheet.cell(row=row[0].row, column=3).value = "-"
                        else:
                            sheet.cell(row=row[0].row, column=3).value = id_value
                    elif code == 'ДМ2':
                        sheet.cell(row=row[0].row, column=3).value = dm2_value
                    elif code == 'КМ2':
                        sheet.cell(row=row[0].row, column=3).value = km2_value
                    elif code == 'YХ2':
                        sheet.cell(row=row[0].row, column=3).value = yx2_value
                    elif code == 'ИД2':
                        if dm2_value == "ниже порога детекции" and km2_value != "ниже порога детекции" and km2_value != "-":
                            sheet.cell(row=row[0].row, column=3).value = "-"
                        else:
                            sheet.cell(row=row[0].row, column=3).value = id2_value
                    elif code == 'ДМ3':
                        sheet.cell(row=row[0].row, column=3).value = dm3_value
                    elif code == 'КМ3':
                        sheet.cell(row=row[0].row, column=3).value = km3_value
                    elif code == 'YХ3':
                        sheet.cell(row=row[0].row, column=3).value = yx3_value
                    elif code == 'ИД3':
                        if dm3_value == "ниже порога детекции" and km3_value != "ниже порога детекции" and km3_value != "-":
                            sheet.cell(row=row[0].row, column=3).value = "-"
                        else:
                            sheet.cell(row=row[0].row, column=3).value = id3_value

                    logging.info(f"Обновлено значение для кода '{code}' на основе однократного поиска.")

                    # Детальный вывод найденных значений
                    found_values, found_samples = search_results
                    num_obrt = len(obrt_search_values)
                    if len(found_values) == 4:
                        # один образец
                        dm, km, yx, id_ = found_values
                        logging.info(f"Найденные значения для {obrt_search_values[0]}: ДМ={dm}, КМ={km}, YХ={yx}, ИД={id_}")
                    elif len(found_values) == 8:
                        # два образца
                        dm1, km1, yx1, id1, dm2, km2, yx2, id2 = found_values
                        logging.info(f"Найденные значения для {obrt_search_values[0]}: ДМ={dm1}, КМ={km1}, YХ={yx1}, ИД={id1}")
                        logging.info(f"Найденные значения для {obrt_search_values[1]}: ДМ={dm2}, КМ={km2}, YХ={yx2}, ИД={id2}")
                    elif len(found_values) == 12:
                        # три образца
                        dm1, km1, yx1, id1, dm2, km2, yx2, id2, dm3, km3, yx3, id3 = found_values
                        logging.info(f"Найденные значения для {obrt_search_values[0]}: ДМ={dm1}, КМ={km1}, YХ={yx1}, ИД={id1}")
                        logging.info(f"Найденные значения для {obrt_search_values[1]}: ДМ={dm2}, КМ={km2}, YХ={yx2}, ИД={id2}")
                        logging.info(f"Найденные значения для {obrt_search_values[2]}: ДМ={dm3}, КМ={km3}, YХ={yx3}, ИД={id3}")
                    else:
                        logging.warning(f"Неожиданная длина результатов поиска: {len(found_values)}")

        except Exception as e:
            raise Exception(f"Ошибка при обновлении значений в Excel: {str(e)}")
        
        # Преобразование ФИО из родительного в именительный падеж (ТОЛЬКО для ЭКЦ)
        try:
            logging.info("Проверка условия для преобразования ФИО: код 1 должен быть 'ЭКЦ' или 'ЭКЦ_нет результата'")
            
            # Проверяем условие для преобразования ФИО
            if code_1_value in ["ЭКЦ", "ЭКЦ_нет результата"] or (code_1_value and isinstance(code_1_value, str) and code_1_value.startswith("ЭКЦ_образец_")):
                logging.info(f"Условие выполнено для преобразования ФИО: код 1 = '{code_1_value}'")
                
                # Переменные для хранения ФИО в родительном падеже
                fr = None
                ir = None
                or_ = None
                
                # Проходим по строкам и ищем ФИО в родительном падеже
                for row in sheet.iter_rows(min_row=2, max_col=3):
                    code = row[0].value
                    value = row[2].value
                    value = strip_spaces(value)  # Удаляем пробелы в начале и конце строки
                    if code == 'ФР':
                        fr = value
                        logging.info(f"Найдено ФИО в родительном падеже (ФР): '{fr}'")
                    elif code == 'ИР':
                        ir = value
                        logging.info(f"Найдено ФИО в родительном падеже (ИР): '{ir}'")
                    elif code == 'ОР':
                        or_ = value
                        logging.info(f"Найдено ФИО в родительном падеже (ОР): '{or_}'")
                
                # Преобразуем и записываем имеющиеся значения
                converted = False
                
                # Преобразуем фамилию, если есть
                if fr:
                    try:
                        fi = convert_to_nominative(fr)
                        # Находим строку для записи фамилии в именительном падеже
                        for row in sheet.iter_rows(min_row=2, max_col=3):
                            if row[0].value == 'ФИ':
                                sheet.cell(row=row[0].row, column=3).value = fi
                                logging.info(f"Записано ФИО в именительном падеже (ФИ): '{fi}'")
                                converted = True
                                break
                    except Exception as e:
                        logging.error(f"Ошибка при преобразовании фамилии '{fr}': {str(e)}")
                
                # Преобразуем имя, если есть
                if ir:
                    try:
                        ii = convert_to_nominative(ir)
                        # Находим строку для записи имени в именительном падеже
                        for row in sheet.iter_rows(min_row=2, max_col=3):
                            if row[0].value == 'ИИ':
                                sheet.cell(row=row[0].row, column=3).value = ii
                                logging.info(f"Записано ФИО в именительном падеже (ИИ): '{ii}'")
                                converted = True
                                break
                    except Exception as e:
                        logging.error(f"Ошибка при преобразовании имени '{ir}': {str(e)}")
                
                # Преобразуем отчество, если есть
                if or_:
                    try:
                        oi = convert_to_nominative(or_)
                        # Находим строку для записи отчества в именительном падеже
                        for row in sheet.iter_rows(min_row=2, max_col=3):
                            if row[0].value == 'ОИ':
                                sheet.cell(row=row[0].row, column=3).value = oi
                                logging.info(f"Записано ФИО в именительном падеже (ОИ): '{oi}'")
                                converted = True
                                break
                    except Exception as e:
                        logging.error(f"Ошибка при преобразовании отчества '{or_}': {str(e)}")
                
                if not converted:
                    logging.info("Не найдено ни одной части ФИО для преобразования")
            else:
                logging.info(f"Преобразование ФИО пропущено, так как код 1 = '{code_1_value}' не соответствует условию")
                
        except Exception as e:
            logging.error(f"Ошибка при преобразовании и записи ФИО: {str(e)}")
            # Продолжаем работу программы несмотря на ошибку

        # Блок для обработки АТО и ИНД
        try:
            logging.info("Обработка строк АТО и ИНД")
            ato_value = None
            ind_row = None
            
            # Ищем строку АТО и запоминаем значение
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                if code == 'АТО':
                    ato_value = value
                    logging.info(f"Найдено значение АТО: '{ato_value}'")
                elif code == 'ИНД':
                    ind_row = row[0].row
                    logging.info(f"Найдена строка ИНД в строке {ind_row}")
            
            # Если нашли и АТО и ИНД
            if ato_value and ind_row:
                # Извлекаем город из адреса
                city = extract_city_from_address(ato_value)
                
                if city:
                    # Путь к файлу с почтовыми индексами
                    postal_codes_path = r"U:\ШАБЛОНЫ\Заключения\СВО\Образцы\postal_codes.xlsx"
                    logging.info(f"Используем файл почтовых индексов: {postal_codes_path}")
                    
                    # Загружаем почтовые индексы
                    postal_codes = load_postal_codes(postal_codes_path)
                    
                    # Ищем индекс для населённого пункта
                    if postal_codes:  # Проверяем, что словарь не пустой
                        postal_code = postal_codes.get(city.lower())
                        if postal_code:
                            logging.info(f"Найден почтовый индекс '{postal_code}' для населённого пункта '{city}'")
                            sheet.cell(row=ind_row, column=3).value = postal_code
                        else:
                            logging.warning(f"Почтовый индекс для населённого пункта '{city}' не найден в файле")
                            # НЕ изменяем значение ячейки - оставляем как есть
                    else:
                        logging.warning("Словарь почтовых индексов пуст, поиск не выполнен")
                else:
                    logging.warning("Не удалось извлечь населённый пункт из адреса АТО")
            else:
                if not ato_value:
                    logging.warning("Не найдено значение АТО")
                if not ind_row:
                    logging.warning("Не найдена строка ИНД")
        except Exception as e:
            logging.error(f"Ошибка при обработке АТО/ИНД: {str(e)}")
            # Продолжаем работу программы несмотря на ошибку

        # Проверка значений в строках ОБРТ, ОБРТ2, ОБРТ3 и очистка соответствующих строк
        try:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                
                # Проверяем, есть ли значение в строках ОБРТ, ОБРТ2, ОБРТ3
                if code in ['ОБРТ', 'ОБРТ2', 'ОБРТ3'] and (value is None or str(value).strip() == ""):
                    # Если значение отсутствует, очищаем соответствующие строки
                    if code == 'ОБРТ':
                        sheet.cell(row=row[0].row, column=3).value = ""  # Очищаем ОБРТ
                        # Очищаем соответствующие строки ДМ, КМ, YХ, ИД
                        for r in sheet.iter_rows(min_row=2, max_col=3):
                            if r[0].value in ['ДМ', 'КМ', 'YХ', 'ИД']:
                                sheet.cell(row=r[0].row, column=3).value = ""
                    elif code == 'ОБРТ2':
                        sheet.cell(row=row[0].row, column=3).value = ""  # Очищаем ОБРТ2
                        # Очищаем соответствующие строки ДМ2, КМ2, YХ2, ИД2
                        for r in sheet.iter_rows(min_row=2, max_col=3):
                            if r[0].value in ['ДМ2', 'КМ2', 'YХ2', 'ИД2']:
                                sheet.cell(row=r[0].row, column=3).value = ""
                    elif code == 'ОБРТ3':
                        sheet.cell(row=row[0].row, column=3).value = ""  # Очищаем ОБРТ3
                        # Очищаем соответствующие строки ДМ3, КМ3, YХ3, ИД3
                        for r in sheet.iter_rows(min_row=2, max_col=3):
                            if r[0].value in ['ДМ3', 'КМ3', 'YХ3', 'ИД3']:
                                sheet.cell(row=r[0].row, column=3).value = ""
        except Exception as e:
            logging.error(f"Ошибка при проверке и очистке строк: {str(e)}")
        
        # Обновление значений для ТЕРИ/ТЕР1/ТЕР2
        for row in sheet.iter_rows(min_row=2, max_col=3):
            code = row[0].value
            if code in ["ТЕРИ", "ТЕР1", "ТЕР2"]:
                if code in ter_results:
                    sheet.cell(row=row[0].row, column=3).value = ter_results[code]
                    logging.info(f"Обновлено значение для кода '{code}': '{ter_results[code]}'")
                else:
                    sheet.cell(row=row[0].row, column=3).value = None
                    logging.info(f"Очищено значение для кода '{code}' (условие не выполнено)")

        # Сохраняем изменения в файле Excel
        try:
            workbook.save(excel_file_path)
            logging.info(f"Изменения сохранены в файле Excel: {excel_file_path}")
        except Exception as e:
            raise Exception(f"Ошибка при сохранении файла Excel: {str(e)}")
        
        # Указываем путь к запускаемой программе
        program_to_run = os.path.normpath(r'F:\Работа\Автоматизация-5\cmd\2_Формирование_заключения_v1.49.1.py')
        logging.info(f"Путь к запускаемой программе: {program_to_run}")
        
        # Проверяем наличие файла
        if not os.path.exists(program_to_run):
            raise FileNotFoundError(f"Файл {program_to_run} не найден.")
        logging.info(f"Файл запускаемой программы найден.")
        
        # Запускаем другую программу
        try:
            subprocess.run(['python', program_to_run])
            logging.info("Программа завершена успешно.")
        except Exception as e:
            raise Exception(f"Ошибка при запуске программы: {str(e)}")
    except Exception as e:
        logging.error(f"Произошла ошибка: {str(e)}")
        show_error_message(f"Произошла ошибка: {str(e)}")

if __name__ == "__main__":
    main()