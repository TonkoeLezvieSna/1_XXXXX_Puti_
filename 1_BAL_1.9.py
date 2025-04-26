#1
#v1.9

# -*- coding: utf-8 -*-
import openpyxl
import os
import xlrd
import subprocess
from pymorphy2 import MorphAnalyzer
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Отображение ошибок
def show_error_message(message):
    try:
        root = tk.Tk()
        root.withdraw()  # Скрываем главное окно Tkinter
        messagebox.showerror("Ошибка", message)
        root.destroy()
    except Exception as e:
        logging.error(f"Ошибка при отображении сообщения об ошибке: {str(e)}")

# Инициализируем морфологический анализатор
morph = MorphAnalyzer()

# Преобразует слово из родительного падежа в именительный падеж.
def convert_to_nominative(word):
    try:
        logging.info(f"Преобразуем слово '{word}' из родительного падежа в именительный падеж.")
        parsed_word = morph.parse(word)[0]
        if parsed_word:
            nominative_form = parsed_word.inflect({'nomn'}).word
            logging.info(f"Преобразованное слово: '{nominative_form.capitalize()}'")
            return nominative_form.capitalize()
        else:
            logging.error(f"Не удалось разобрать слово: '{word}'")
            return word
    except Exception as e:
        logging.error(f"Ошибка при преобразовании слова '{word}': {str(e)}")
        return word

# Удаление пробелов в начале и конце строки
def strip_spaces(value):
    if value and isinstance(value, str):
        return value.strip()
    return value

# Заменяет точку на запятую в числовых значениях
def replace_dot_with_comma(value):
    if isinstance(value, (int, float)):
        value_str = str(value)
        if ',' not in value_str:
            return value_str.replace('.', ',')
    elif isinstance(value, str) and '.' in value and value.replace('.', '').isdigit():
        return value.replace('.', ',')
    return str(value) if value is not None else value

# Функция для поиска значений РТ в папке
def find_values_in_excel_files(folder_path, obrt_values):
    # Инициализация значений по умолчанию
    default_values = {
        "dm": "ниже порога детекции",
        "km": "ниже порога детекции",
        "yx": "ниже порога детекции",
        "id": "-",
    }

    # Создаём словарь для результатов
    results = {}
    for obrt in obrt_values:
        results[obrt] = default_values.copy()

    # Рекурсивный обход файлов
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            
            # Пропускаем не-Excel файлы
            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                continue

            try:
                # Открываем файл (.xlsx или .xls)
                if file_name.endswith('.xlsx'):
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook.active
                    rows = list(sheet.iter_rows(min_row=2, max_col=10, values_only=True))
                else:  # .xls
                    workbook = xlrd.open_workbook(file_path)
                    sheet = workbook.sheet_by_index(0)
                    rows = []
                    for row_idx in range(1, sheet.nrows):
                        rows.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])

                # Обработка строк
                for row in rows:
                    if len(row) < 3:  # Проверяем, что есть Well, Sample Name и Target Name
                        continue

                    sample_name = strip_spaces(str(row[1]))
                    if sample_name not in obrt_values:
                        continue

                    row_type = strip_spaces(str(row[2]))
                    if not row_type:
                        continue

                    # Парсинг числовых значений
                    def parse_value(value, is_id=False):
                        if value is None:
                            return None
                        value = strip_spaces(str(value))
                        if value.replace(".", "", 1).isdigit():
                            return round(float(value), 2 if is_id else 4)
                        return None

                    # Для T.Large Autosomal (ДМ и ИД)
                    if row_type == "T.Large Autosomal":
                        dm_value = parse_value(row[4] if len(row) > 4 else None)
                        id_value = parse_value(row[5] if len(row) > 5 else None, is_id=True)
                        
                        if dm_value is not None:
                            results[sample_name]["dm"] = replace_dot_with_comma(dm_value)
                        if id_value is not None:
                            results[sample_name]["id"] = replace_dot_with_comma(id_value if id_value >= 1 else 0)

                    # Для T.Small Autosomal (КМ)
                    elif row_type == "T.Small Autosomal":
                        km_value = parse_value(row[4] if len(row) > 4 else None)
                        if km_value is not None:
                            results[sample_name]["km"] = replace_dot_with_comma(km_value)

                    # Для T.Y (YХ)
                    elif row_type == "T.Y":
                        yx_value = parse_value(row[4] if len(row) > 4 else None)
                        if yx_value is not None:
                            results[sample_name]["yx"] = replace_dot_with_comma(yx_value)

            except Exception as e:
                logging.error(f"Ошибка при обработке файла {file_name}: {str(e)}")

    # Возвращаем результаты
    output = []
    for obrt in obrt_values:
        output.extend([
            results[obrt]["dm"],
            results[obrt]["km"],
            results[obrt]["yx"],
            results[obrt]["id"],
        ])
    return tuple(output)
    
# Главная функция
def main():
    try:
        # Получаем текущую директорию
        current_directory = os.path.dirname(os.path.abspath(__file__))
        logging.info(f"Текущая директория: {current_directory}")
        
        # Находим все файлы Excel в текущей директории
        excel_files = [f for f in os.listdir(current_directory) if f.endswith('.xlsx')]
        logging.info(f"Найденные файлы Excel: {excel_files}")
        
        # Проверяем, что файл Excel существует и единственный
        if not excel_files:
            raise FileNotFoundError("В текущей директории не найдено файлов Excel.")
        if len(excel_files) != 1:
            raise FileNotFoundError("В текущей директории должен быть только один файл Excel.")
        logging.info(f"Файл Excel найден: {excel_files[0]}")
        
        # Полный путь к файлу Excel
        excel_file_path = os.path.join(current_directory, excel_files[0])
        logging.info(f"Полный путь к файлу Excel: {excel_file_path}")
        
        # Открываем файл Excel
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active
            logging.info(f"Файл Excel открыт успешно.")
        except Exception as e:
            raise Exception(f"Ошибка при открытии файла Excel: {str(e)}")
        
        # Проверка условия для кода 1
        code_1_value = None
        obrt_values = []
        obrt2_value = None
        obrt3_value = None

        # Собираем значения для кода 1 и ОБРТ, ОБРТ2, ОБРТ3
        for row in sheet.iter_rows(min_row=2, max_col=3):
            code = row[0].value
            value = row[2].value
            if code == 1:
                code_1_value = str(value).strip() if value else None
            elif code == 'ОБРТ':
                obrt_value = str(value).strip() if value else None
                if obrt_value:
                    obrt_values.append(obrt_value)
            elif code == 'ОБРТ2':
                obrt2_value = str(value).strip() if value else None
                if obrt2_value:
                    obrt_values.append(obrt2_value)
            elif code == 'ОБРТ3':
                obrt3_value = str(value).strip() if value else None
                if obrt3_value:
                    obrt_values.append(obrt3_value)

        # Проверяем условие для завершения программы
        svo_condition = code_1_value in ["Кость СВО", "Кость СВО нет пригодной ДНК"]

        # Проверка наличия значений ОБРТ
        if not any(obrt_values) and svo_condition:
            error_msg = "Ошибка: В графе с кодом 1 указано 'Кость СВО' или 'Кость СВО нет пригодной ДНК', но значения ОБРТ отсутствуют"
            logging.error(error_msg)
            raise ValueError(error_msg)

        # Проверяем наличие значений в файлах Excel для всех указанных ОБРТ
        try:
            if obrt_values:
                found_values = find_values_in_excel_files(r"F:\Работа\Автоматизация-5\2 - РТ", obrt_values)
                
                # Проверка для ОБРТ
                if 'ОБРТ' in [code[0].value for code in sheet.iter_rows(min_row=2, max_col=1) if code[0].value]:
                    if found_values[0] == "ниже порога детекции" and found_values[1] == "ниже порога детекции" and found_values[2] == "ниже порога детекции" and found_values[3] == "-":
                        error_msg = f"Значения для ОБРТ '{obrt_values[0]}' не найдены в файлах Excel"
                        logging.error(error_msg)
                        if svo_condition:
                            raise ValueError(error_msg)
                
                # Проверка для ОБРТ2
                if 'ОБРТ2' in [code[0].value for code in sheet.iter_rows(min_row=2, max_col=1) if code[0].value] and len(obrt_values) > 1:
                    if len(found_values) >= 8 and found_values[4] == "ниже порога детекции" and found_values[5] == "ниже порога детекции" and found_values[6] == "ниже порога детекции" and found_values[7] == "-":
                        error_msg = f"Значения для ОБРТ2 '{obrt_values[1]}' не найдены в файлах Excel"
                        logging.error(error_msg)
                        if svo_condition:
                            raise ValueError(error_msg)
                
                # Проверка для ОБРТ3
                if 'ОБРТ3' in [code[0].value for code in sheet.iter_rows(min_row=2, max_col=1) if code[0].value] and len(obrt_values) > 2:
                    if len(found_values) >= 12 and found_values[8] == "ниже порога детекции" and found_values[9] == "ниже порога детекции" and found_values[10] == "ниже порога детекции" and found_values[11] == "-":
                        error_msg = f"Значения для ОБРТ3 '{obrt_values[2]}' не найдены в файлах Excel"
                        logging.error(error_msg)
                        if svo_condition:
                            raise ValueError(error_msg)
        except Exception as e:
            if svo_condition:
                raise Exception(f"Ошибка при поиске значений для ОБРТ: {str(e)}")
            else:
                logging.error(f"Ошибка при поиске значений для ОБРТ: {str(e)}")
                    
            
        # Новые значения, которые нужно вставить в столбец 3
        new_values = {
            '1СТ': 'Бевза Алексей Леонидович',
            '2СТ': 'образование высшее медицинское, специальность – судебно-медицинская экспертиза,',
            '3СТ': 'стаж работы по специальности с 1.09.2004 г., высшая квалификационная категория,',
            '4СТ': 'кандидат медицинских наук, врач судебно-медицинский эксперт',
            '5СТ': ' ',
            '6СТ': 'произвёл',
            '1ИЭ': 'А.Л. Бевза',
            '2ИЭ': '                                                     Бевза А.Л.',
            '3ИЭ': 'Бевза Алексей Леонидович',
            '4ИЭ': 'врач судебно-медицинский эксперт ',
            '5ИЭ': 'отделения судебно-биологической',
            '6ИЭ': 'и генетической экспертизы, к.м.н.',
            '7ИЭ': 'тел. +7 (3452) 49-43-98 (доб. 1345)',
            2: 'F:\\Работа\\Автоматизация-5\\2',  # Вставьте новый путь для кода 2
            3: 'F:\\Работа\\Автоматизация-5\\3',  # Вставьте новый путь для кода 3
            4: 'F:\\Работа\\Автоматизация-5\\4',   # Вставьте новый путь для кода 4
            5: 'F:\\Работа\\Автоматизация-5\\5',  # Вставьте новый путь для кода 5
            6: 'F:\\Работа\\Автоматизация-5\\6',   # Вставьте новый путь для кода 6
        }
        logging.info(f"Новые значения для вставки: {new_values}")
        
        # Проходим по строкам и обновляем значения в столбец 3 для заданных кодов
        try:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                
                # Если код находится в new_values, обновляем значение
                if code in new_values:
                    sheet.cell(row=row[0].row, column=3).value = new_values[code]
                    logging.info(f"Обновлено значение для кода '{code}': '{new_values[code]}'")
                
                # Если код — ДМ, КМ, YХ, ИД, ДМ2, КМ2, YХ2, ИД2, ДМ3, КМ3, YХ3, ИД3, ищем значения в файлах Excel
                elif code in ['ДМ', 'КМ', 'YХ', 'ИД', 'ДМ2', 'КМ2', 'YХ2', 'ИД2', 'ДМ3', 'КМ3', 'YХ3', 'ИД3']:
                    # Ищем значение ОБРТ, ОБРТ2, ОБРТ3 в Карте
                    obrt_values = []
                    for r in sheet.iter_rows(min_row=2, max_col=3):
                        if r[0].value in ['ОБРТ', 'ОБРТ2', 'ОБРТ3'] and r[2].value and str(r[2].value).strip():
                            obrt_values.append(r[2].value)
                    
                    # Если ОБРТ, ОБРТ2, ОБРТ3 найдены, ищем значения в файлах Excel
                    if obrt_values:
                        try:
                            values = find_values_in_excel_files(r"F:\Работа\Автоматизация-5\2 - РТ", obrt_values)
                            logging.info(f"Найдены raw-значения: {values}")

                            # Обработка результатов
                            if len(values) == 12:  # Полный набор (ОБРТ, ОБРТ2, ОБРТ3)
                                dm_value, km_value, yx_value, id_value, dm2_value, km2_value, yx2_value, id2_value, dm3_value, km3_value, yx3_value, id3_value = values
                                logging.info(f"""
Обновление значений:
[ОБРТ {obrt_values[0]}]
ДМ = {dm_value}, КМ = {km_value}, YХ = {yx_value}, ИД = {id_value}
[ОБРТ2 {obrt_values[1]}]
ДМ2 = {dm2_value}, КМ2 = {km2_value}, YХ2 = {yx2_value}, ИД2 = {id2_value}
[ОБРТ3 {obrt_values[2]}]
ДМ3 = {dm3_value}, КМ3 = {km3_value}, YХ3 = {yx3_value}, ИД3 = {id3_value}
""")
                            elif len(values) == 8:  # Только ОБРТ и ОБРТ2
                                dm_value, km_value, yx_value, id_value, dm2_value, km2_value, yx2_value, id2_value = values
                                dm3_value, km3_value, yx3_value, id3_value = "ниже порога детекции", "ниже порога детекции", "ниже порога детекции", "-"
                                logging.info(f"""
Обновление значений (ОБРТ3 не найден):
[ОБРТ {obrt_values[0]}]
ДМ = {dm_value}, КМ = {km_value}, YХ = {yx_value}, ИД = {id_value}
[ОБРТ2 {obrt_values[1]}]
ДМ2 = {dm2_value}, КМ2 = {km2_value}, YХ2 = {yx2_value}, ИД2 = {id2_value}
[ОБРТ3]
Установлены значения по умолчанию
""")
                            elif len(values) == 4:  # Только ОБРТ
                                dm_value, km_value, yx_value, id_value = values
                                dm2_value, km2_value, yx2_value, id2_value = "ниже порога детекции", "ниже порога детекции", "ниже порога детекции", "-"
                                dm3_value, km3_value, yx3_value, id3_value = "ниже порога детекции", "ниже порога детекции", "ниже порога детекции", "-"
                                logging.info(f"""
Обновление значений (только ОБРТ найден):
[ОБРТ {obrt_values[0]}]
ДМ = {dm_value}, КМ = {km_value}, YХ = {yx_value}, ИД = {id_value}
[ОБРТ2/ОБРТ3]
Установлены значения по умолчанию
""")
                            else:
                                logging.error(f"Ошибка: получено {len(values)} значений, ожидалось 4, 8 или 12")
                                raise ValueError(f"Непредвиденное количество значений: {len(values)}")        
                                
                            # Проверяем, найдены ли значения
                            if (dm_value == "ниже порога детекции" and km_value == "ниже порога детекции" and yx_value == "ниже порога детекции" and id_value == "-" and
                                dm2_value == "ниже порога детекции" and km2_value == "ниже порога детекции" and yx2_value == "ниже порога детекции" and id2_value == "-" and
                                dm3_value == "ниже порога детекции" and km3_value == "ниже порога детекции" and yx3_value == "ниже порога детекции" and id3_value == "-"):
                                raise ValueError(f"Значения ОБРТ, ОБРТ2, ОБРТ3 '{obrt_values}' не найдены в файлах Excel.")
        
                            # Заполняем ячейки Карты
                            if code == 'ДМ':
                                sheet.cell(row=row[0].row, column=3).value = dm_value
                            elif code == 'КМ':
                                sheet.cell(row=row[0].row, column=3).value = km_value
                            elif code == 'YХ':
                                sheet.cell(row=row[0].row, column=3).value = yx_value
                            elif code == 'ИД':
                                # Проверяем условие: если ДМ = "ниже порога детекции" и КМ - числовое значение, ставим "-"
                                if dm_value == "ниже порога детекции" and km_value != "ниже порога детекции" and km_value != "-":
                                    sheet.cell(row=row[0].row, column=3).value = "-"
                                else:
                                    sheet.cell(row=row[0].row, column=3).value = id_value
                            elif code == 'ДМ2':
                                sheet.cell(row=row[0].row, column=3).value = dm2_value
                            elif code == 'КМ2':
                                sheet.cell(row=row[0].row, column=3).value = km2_value
                            elif code == 'YХ2':
                                sheet.cell(row=row[0].row, column=3).value = yx2_value
                            elif code == 'ИД2':
                                # Проверяем условие для ИД2
                                if dm2_value == "ниже порога детекции" and km2_value != "ниже порога детекции" and km2_value != "-":
                                    sheet.cell(row=row[0].row, column=3).value = "-"
                                else:
                                    sheet.cell(row=row[0].row, column=3).value = id2_value
                            elif code == 'ДМ3':
                                sheet.cell(row=row[0].row, column=3).value = dm3_value
                            elif code == 'КМ3':
                                sheet.cell(row=row[0].row, column=3).value = km3_value
                            elif code == 'YХ3':
                                sheet.cell(row=row[0].row, column=3).value = yx3_value
                            elif code == 'ИД3':
                                # Проверяем условие для ИД3
                                if dm3_value == "ниже порога детекции" and km3_value != "ниже порога детекции" and km3_value != "-":
                                    sheet.cell(row=row[0].row, column=3).value = "-"
                                else:
                                    sheet.cell(row=row[0].row, column=3).value = id3_value        
        
                            logging.info(f"Обновлено значение для кода '{code}' на основе ОБРТ: {obrt_values}")
                        except Exception as e:
                            logging.error(f"Ошибка при поиске значений для ОБРТ '{obrt_values}': {str(e)}")
                            show_error_message(f"Ошибка: {str(e)}")
                            sheet.cell(row=row[0].row, column=3).value = "ниже порога детекции"
                    else:
                        logging.warning(f"Не найдены значения ОБРТ, ОБРТ2, ОБРТ3 в Карте.")
                        sheet.cell(row=row[0].row, column=3).value = "ниже порога детекции"

        except Exception as e:
            raise Exception(f"Ошибка при обновлении значений в Excel: {str(e)}")
        
        # Переменные для хранения ФИО в родительном падеже
        fr = None
        ir = None
        or_ = None
        
        # Проходим по строкам и ищем ФИО в родительном падеже
        try:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                value = strip_spaces(value)  # Удаляем пробелы в начале и конце строки
                if code == 'ФР':
                    fr = value
                    logging.info(f"Найдено ФИО в родительном падеже (ФР): '{fr}'")
                elif code == 'ИР':
                    ir = value
                    logging.info(f"Найдено ФИО в родительном падеже (ИР): '{ir}'")
                elif code == 'ОР':
                    or_ = value
                    logging.info(f"Найдено ФИО в родительном падеже (ОР): '{or_}'")
        except Exception as e:
            raise Exception(f"Ошибка при поиске ФИО в родительном падеже: {str(e)}")
        
        # Если все части ФИО в родительном падеже найдены, преобразуем их
        if fr and ir and or_:
            try:
                fi = convert_to_nominative(fr)
                ii = convert_to_nominative(ir)
                oi = convert_to_nominative(or_)
                
                # Находим строки для записи ФИО в именительном падеже
                for row in sheet.iter_rows(min_row=2, max_col=3):
                    code = row[0].value
                    if code == 'ФИ':
                        sheet.cell(row=row[0].row, column=3).value = fi
                        logging.info(f"Записано ФИО в именительном падеже (ФИ): '{fi}'")
                    elif code == 'ИИ':
                        sheet.cell(row=row[0].row, column=3).value = ii
                        logging.info(f"Записано ФИО в именительном падеже (ИИ): '{ii}'")
                    elif code == 'ОИ':
                        sheet.cell(row=row[0].row, column=3).value = oi
                        logging.info(f"Записано ФИО в именительном падеже (ОИ): '{oi}'")
            except Exception as e:
                raise Exception(f"Ошибка при преобразовании и записи ФИО: {str(e)}")
        else:
            logging.info("Не все части ФИО в родительном падеже найдены. Пропускаем преобразование.")

        # Проверка значений в строках ОБРТ, ОБРТ2, ОБРТ3 и очистка соответствующих строк
        try:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                
                # Проверяем, есть ли значение в строках ОБРТ, ОБРТ2, ОБРТ3
                if code in ['ОБРТ', 'ОБРТ2', 'ОБРТ3'] and (value is None or str(value).strip() == ""):
                    # Если значение отсутствует, очищаем соответствующие строки
                    if code == 'ОБРТ':
                        sheet.cell(row=row[0].row, column=3).value = ""  # Очищаем ОБРТ
                        # Очищаем соответствующие строки ДМ, КМ, YХ, ИД
                        for r in sheet.iter_rows(min_row=2, max_col=3):
                            if r[0].value in ['ДМ', 'КМ', 'YХ', 'ИД']:
                                sheet.cell(row=r[0].row, column=3).value = ""
                    elif code == 'ОБРТ2':
                        sheet.cell(row=row[0].row, column=3).value = ""  # Очищаем ОБРТ2
                        # Очищаем соответствующие строки ДМ2, КМ2, YХ2, ИД2
                        for r in sheet.iter_rows(min_row=2, max_col=3):
                            if r[0].value in ['ДМ2', 'КМ2', 'YХ2', 'ИД2']:
                                sheet.cell(row=r[0].row, column=3).value = ""
                    elif code == 'ОБРТ3':
                        sheet.cell(row=row[0].row, column=3).value = ""  # Очищаем ОБРТ3
                        # Очищаем соответствующие строки ДМ3, КМ3, YХ3, ИД3
                        for r in sheet.iter_rows(min_row=2, max_col=3):
                            if r[0].value in ['ДМ3', 'КМ3', 'YХ3', 'ИД3']:
                                sheet.cell(row=r[0].row, column=3).value = ""
        except Exception as e:
            logging.error(f"Ошибка при проверке и очистке строк: {str(e)}")
        
        # Сохраняем изменения в файле Excel
        try:
            workbook.save(excel_file_path)
            logging.info(f"Изменения сохранены в файле Excel: {excel_file_path}")
        except Exception as e:
            raise Exception(f"Ошибка при сохранении файла Excel: {str(e)}")
        
        # Указываем путь к запускаемой программе
        program_to_run = os.path.normpath('F:\\Работа\\Автоматизация-5\\cmd\\2_v1.31.py')
        logging.info(f"Путь к запускаемой программе: {program_to_run}")
        
        # Проверяем наличие файла
        if not os.path.exists(program_to_run):
            raise FileNotFoundError(f"Файл {program_to_run} не найден.")
        logging.info(f"Файл запускаемой программы найден.")
        
        # Запускаем другую программу
        try:
            subprocess.run(['python', program_to_run])
            logging.info("Программа завершена успешно.")
        except Exception as e:
            raise Exception(f"Ошибка при запуске программы: {str(e)}")
    except Exception as e:
        logging.error(f"Произошла ошибка: {str(e)}")
        show_error_message(f"Произошла ошибка: {str(e)}")

if __name__ == "__main__":
    main()
