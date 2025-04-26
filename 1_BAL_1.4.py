#1
#v1.4

# -*- coding: utf-8 -*-
import openpyxl
import os
import subprocess
from pymorphy2 import MorphAnalyzer
import tkinter as tk
from tkinter import messagebox
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Инициализируем морфологический анализатор
morph = MorphAnalyzer()

# Преобразует слово из родительного падежа в именительный падеж.
def convert_to_nominative(word):
    try:
        logging.info(f"Преобразуем слово '{word}' из родительного падежа в именительный падеж.")
        parsed_word = morph.parse(word)[0]
        if parsed_word:
            nominative_form = parsed_word.inflect({'nomn'}).word
            logging.info(f"Преобразованное слово: '{nominative_form.capitalize()}'")
            return nominative_form.capitalize()
        else:
            logging.error(f"Не удалось разобрать слово: '{word}'")
            return word
    except Exception as e:
        logging.error(f"Ошибка при преобразовании слова '{word}': {str(e)}")
        return word

# Отображение ошибок
def show_error_message(message):
    try:
        root = tk.Tk()
        root.withdraw()  # Скрываем главное окно Tkinter
        messagebox.showerror("Ошибка", message)
        root.destroy()
    except Exception as e:
        logging.error(f"Ошибка при отображении сообщения об ошибке: {str(e)}")

# Главная функция
def main():
    try:
        # Получаем текущую директорию
        current_directory = os.path.dirname(os.path.abspath(__file__))
        logging.info(f"Текущая директория: {current_directory}")
        
        # Находим все файлы Excel в текущей директории
        excel_files = [f for f in os.listdir(current_directory) if f.endswith('.xlsx')]
        logging.info(f"Найденные файлы Excel: {excel_files}")
        
        # Проверяем, что файл Excel существует и единственный
        if not excel_files:
            raise FileNotFoundError("В текущей директории не найдено файлов Excel.")
        if len(excel_files) != 1:
            raise FileNotFoundError("В текущей директории должен быть только один файл Excel.")
        logging.info(f"Файл Excel найден: {excel_files[0]}")
        
        # Полный путь к файлу Excel
        excel_file_path = os.path.join(current_directory, excel_files[0])
        logging.info(f"Полный путь к файлу Excel: {excel_file_path}")
        
        # Открываем файл Excel
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active
            logging.info(f"Файл Excel открыт успешно.")
        except Exception as e:
            raise Exception(f"Ошибка при открытии файла Excel: {str(e)}")
        
        # Новые значения, которые нужно вставить в столбец 3
        new_values = {
            '1СТ': 'Бевза Алексей Леонидович',
            '2СТ': 'образование высшее медицинское, специальность – судебно-медицинская экспертиза,',
            '3СТ': 'стаж работы по специальности с 1.09.2004 г., высшая квалификационная категория,',
            '4СТ': 'кандидат медицинских наук, врач судебно-медицинский эксперт',
            '5СТ': ' ',
            '6СТ': 'произвёл',
            '1ИЭ': 'А.Л. Бевза',
            '2ИЭ': '                                                     Бевза А.Л.',
            '3ИЭ': 'Бевза Алексей Леонидович',
            '4ИЭ': 'врач судебно-медицинский эксперт ',
            '5ИЭ': 'отделения судебно-биологической',
            '6ИЭ': 'и генетической экспертизы, к.м.н.',
            '7ИЭ': 'тел. +7 (3452) 49-43-98 (доб. 1345)',
            2: 'U:\\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\\Беуза\\Экспертизы\\2025\\Уголовные',  # Вставьте новый путь для кода 2
            3: 'U:\\ШАБЛОНЫ\\Заключения\\СВО\\Образцы',  # Вставьте новый путь для кода 3
            4: 'U:\\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\\Беуза\\ANALIZATOR',   # Вставьте новый путь для кода 4
            5: 'U:\\Архив\\Архив 2025',  # Вставьте новый путь для кода 5
            6: 'U:\\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\\Ростов СВО',   # Вставьте новый путь для кода 6
        }
        logging.info(f"Новые значения для вставки: {new_values}")
        
        # Проходим по строкам и обновляем значения в столбце 3 для заданных кодов
        try:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                if code in new_values:
                    sheet.cell(row=row[0].row, column=3).value = new_values[code]
                    logging.info(f"Обновлено значение для кода '{code}': '{new_values[code]}'")
        except Exception as e:
            raise Exception(f"Ошибка при обновлении значений в Excel: {str(e)}")
        
        # Переменные для хранения ФИО в родительном падеже
        fr = None
        ir = None
        or_ = None
        
        # Проходим по строкам и ищем ФИО в родительном падеже
        try:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                code = row[0].value
                value = row[2].value
                if code == 'ФР':
                    fr = value
                    logging.info(f"Найдено ФИО в родительном падеже (ФР): '{fr}'")
                elif code == 'ИР':
                    ir = value
                    logging.info(f"Найдено ФИО в родительном падеже (ИР): '{ir}'")
                elif code == 'ОР':
                    or_ = value
                    logging.info(f"Найдено ФИО в родительном падеже (ОР): '{or_}'")
        except Exception as e:
            raise Exception(f"Ошибка при поиске ФИО в родительном падеже: {str(e)}")
        
        # Если все части ФИО в родительном падеже найдены, преобразуем их
        if fr and ir and or_:
            try:
                fi = convert_to_nominative(fr)
                ii = convert_to_nominative(ir)
                oi = convert_to_nominative(or_)
                
                # Находим строки для записи ФИО в именительном падеже
                for row in sheet.iter_rows(min_row=2, max_col=3):
                    code = row[0].value
                    if code == 'ФИ':
                        sheet.cell(row=row[0].row, column=3).value = fi
                        logging.info(f"Записано ФИО в именительном падеже (ФИ): '{fi}'")
                    elif code == 'ИИ':
                        sheet.cell(row=row[0].row, column=3).value = ii
                        logging.info(f"Записано ФИО в именительном падеже (ИИ): '{ii}'")
                    elif code == 'ОИ':
                        sheet.cell(row=row[0].row, column=3).value = oi
                        logging.info(f"Записано ФИО в именительном падеже (ОИ): '{oi}'")
            except Exception as e:
                raise Exception(f"Ошибка при преобразовании и записи ФИО: {str(e)}")
        
        # Сохраняем изменения в файле Excel
        try:
            workbook.save(excel_file_path)
            logging.info(f"Изменения сохранены в файле Excel: {excel_file_path}")
        except Exception as e:
            raise Exception(f"Ошибка при сохранении файла Excel: {str(e)}")
        
        # Указываем путь к запускаемой программе
        program_to_run = os.path.normpath('F:\\Работа\\Автоматизация-5\\cmd\\V1.15.py')
        logging.info(f"Путь к запускаемой программе: {program_to_run}")
        
        # Проверяем наличие файла
        if not os.path.exists(program_to_run):
            raise FileNotFoundError(f"Файл {program_to_run} не найден.")
        logging.info(f"Файл запускаемой программы найден.")
        
        # Запускаем другую программу
        try:
            subprocess.run(['python', program_to_run])
            logging.info("Программа завершена успешно.")
        except Exception as e:
            raise Exception(f"Ошибка при запуске программы: {str(e)}")
    except Exception as e:
        logging.error(f"Произошла ошибка: {str(e)}")
        show_error_message(f"Произошла ошибка: {str(e)}")

if __name__ == "__main__":
    main()
