# -*- coding: utf-8 -*-
import openpyxl
import os
import subprocess
from pymorphy2 import MorphAnalyzer

# Инициализируем морфологический анализатор
morph = MorphAnalyzer()

def convert_to_nominative(word):
    """
    Преобразует слово из родительного падежа в именительный падеж.
    """
    parsed_word = morph.parse(word)[0]
    nominative_form = parsed_word.inflect({'nomn'}).word
    return nominative_form.capitalize()

# Получаем текущую директорию
current_directory = os.path.dirname(os.path.abspath(__file__))
# Находим все файлы Excel в текущей директории
excel_files = [f for f in os.listdir(current_directory) if f.endswith('.xlsx')]
# Проверяем, что файл Excel существует и единственный
if len(excel_files) != 1:
    raise FileNotFoundError("В текущей директории должен быть только один файл Excel.")
# Полный путь к файлу Excel
excel_file_path = os.path.join(current_directory, excel_files[0])
# Открываем файл Excel
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Новые значения, которые нужно вставить в столбец 3
new_values = {
    '1СТ': 'Бевза Алексей Леонидович',  # Вставьте новое значение для кода 1СТ
    '2СТ': 'образование высшее медицинское, специальность – судебно-медицинская экспертиза,',  # Вставьте новое значение для кода 2СТ
    '3СТ': 'стаж работы по специальности с 1.09.2004 г., высшая квалификационная категория,',  # Вставьте новое значение для кода 3СТ
    '4СТ': 'кандидат медицинских наук, врач судебно-медицинский эксперт',  # Вставьте новое значение для кода 4СТ
    '5СТ': ' ',  # Вставьте новое значение для кода 5СТ
    '6СТ': 'произвёл',  # Вставьте новое значение для кода 5СТ
    '1ИЭ': 'А.Л. Бевза',  # Вставьте новое значение для кода 1ИЭ
    '2ИЭ': '                                                     Бевза А.Л.',  # Вставьте новое значение для кода 2ИЭ
    '3ИЭ': 'Бевза Алексей Леонидович',  # Вставьте новое значение для кода 3ИЭ
    '4ИЭ': 'врач судебно-медицинский эксперт ',  # Вставьте новое значение для кода 4ИЭ
    '5ИЭ': 'отделения судебно-биологической',  # Вставьте новое значение для кода 5ИЭ
    '6ИЭ': 'и генетической экспертизы, к.м.н.',  # Вставьте новое значение для кода 6ИЭ
    '7ИЭ': 'тел. +7 (3452) 49-43-98 (доб. 1345)',  # Вставьте новое значение для кода 7ИЭ
    2: 'F:\\Работа\\Автоматизация-5\\2',  # Вставьте новый путь для кода 2
    3: 'F:\\Работа\\Автоматизация-5\\3',  # Вставьте новый путь для кода 3
    4: 'F:\\Работа\\Автоматизация-5\\4',   # Вставьте новый путь для кода 4
    5: 'F:\\Работа\\Автоматизация-5\\5',  # Вставьте новый путь для кода 5
    6: 'F:\\Работа\\Автоматизация-5\\6',   # Вставьте новый путь для кода 6
}

# Проходим по строкам и обновляем значения в столбце 3 для заданных кодов
for row in sheet.iter_rows(min_row=2, max_col=3):  # Начинаем с второй строки (первая - заголовок)
    code = row[0].value
    value = row[2].value  # Значение находится в третьей колонке
    if code in new_values:
        sheet.cell(row=row[0].row, column=3).value = new_values[code]

# Переменные для хранения ФИО в родительном падеже
fr = None
ir = None
or_ = None

# Проходим по строкам и ищем ФИО в родительном падеже
for row in sheet.iter_rows(min_row=2, max_col=3):  # Начинаем с второй строки (первая - заголовок)
    code = row[0].value
    value = row[2].value  # Значение находится в третьей колонке
    if code == 'ФР':
        fr = value
    elif code == 'ИР':
        ir = value
    elif code == 'ОР':
        or_ = value

# Если все части ФИО в родительном падеже найдены, преобразуем их
if fr and ir and or_:
    fi = convert_to_nominative(fr)
    ii = convert_to_nominative(ir)
    oi = convert_to_nominative(or_)
    
    # Находим строки для записи ФИО в именительном падеже
    for row in sheet.iter_rows(min_row=2, max_col=3):  # Начинаем с второй строки (первая - заголовок)
        code = row[0].value
        if code == 'ФИ':
            sheet.cell(row=row[0].row, column=3).value = fi  # ФИ в именительном падеже
        elif code == 'ИИ':
            sheet.cell(row=row[0].row, column=3).value = ii  # ИИ в именительном падеже
        elif code == 'ОИ':
            sheet.cell(row=row[0].row, column=3).value = oi  # ОИ в именительном падеже

# Сохраняем изменения в файле Excel
workbook.save(excel_file_path)

# Указываем путь к запускаемой программе
program_to_run = os.path.normpath('F:\\Работа\\Автоматизация-5\\cmd\\V1.10.py')
# Проверяем наличие файла
if not os.path.exists(program_to_run):
    raise FileNotFoundError(f"Файл {program_to_run} не найден.")
# Запускаем другую программу
subprocess.run(['python', program_to_run])
